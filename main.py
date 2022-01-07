# speech recognition library import

import speech_recognition as sr

# openpyxl library import

import openpyxl

# two objects created for Recognizer module in Speech Recognition library

r1 = sr.Recognizer()
r2 = sr.Recognizer()

# Microphone module used to capture audio using a mic in the system

with sr.Microphone() as source:
    print("Speak now!")

    # capturing audio in audio variable using listen module from SR

    audio = r1.listen(source)
    try:

        # converted audio to text and stored in text variable using recognize_google module from SR

        text = r2.recognize_google(audio)
    except sr.UnknownValueError:
        print('error')
    except sr.RequestError as e:
        print("failed".format(e))

# text is being spit into a list

words = text.split()
words[0] = int(words[0])

# path for the excel file is declared

path = "E:\\Py Projects\\VoWa\\VoWa.xlsx"

# wb variable declared to contain the declared excel

wb = openpyxl.load_workbook(path)

# to monitor or enter the to the active sheet in the opened excel

sheet = wb.active

# loop to make 1000 entries in the excel without being overwritten

for i in range(1, 1000):

    # to store the first element of words list in empty cell orderly

    if sheet.cell(row=i, column=1).value is None:
        c1 = sheet.cell(row=i, column=1)
        c1.value = words[0]

    # to store the last element of words list in empty cell orderly

    if sheet.cell(row=i, column=2).value is None:
        c2 = sheet.cell(row=i, column=2)
        c2.value = words[-1]
        break

sheet['D1'] = 'Total Expenditure'

# to display the sum of expenditure inputs on column A

sheet['E1'] = '=sum(A1:A1000)'

# to save the changes done to the excel

wb.save(path)
